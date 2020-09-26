###### functions for easy system data input and output in Excel

import pandas as pd
import numpy as np
import tkinter
import tkinter.filedialog
from openpyxl import load_workbook

#filename = "System_Input.xlsx"

def select_file():
    root = tkinter.Tk()
    filename = tkinter.filedialog.askopenfilename(
            title="Open source file",
            filetypes = (("Excel files","*.xlsx"),("all files","*.*")))
    root.destroy()
    return filename

def read_source(sheet,
                num_rows,
                num_cols,
                source_file,
                rows_to_skip=0):

    if rows_to_skip == 0:
        skip = []
    else:
        skip = range(rows_to_skip)
    
    df = pd.read_excel(source_file,
                       sheet_name = sheet,
                       index_col = 0,
                       nrows = num_rows,
                       usecols = range(num_cols+1),
                       skiprows = skip)

    # convert dataframe to numpy array
    var = df.to_numpy()
    var = np.nan_to_num(var, copy=False, nan=-1)
    
    return var

def write_output(sheet_name,
                 array,
                 idx,
                 cols,
                 output_file,
                 rows_to_skip=0,
                 float_format="%.4f"):
    """

    """
    data_df = pd.DataFrame(data = array,
                           index = idx,
                           columns = cols)
    writer = pd.ExcelWriter(output_file, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(output_file)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if rows_to_skip == 0 and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        print("File not found")

    # write out the new sheet
    data_df.to_excel(writer, sheet_name, startrow=rows_to_skip,
                     float_format=float_format)

    # save the workbook
    writer.save()    
    
    


        

    
