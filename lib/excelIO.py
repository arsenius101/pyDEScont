import pandas as pd
import numpy as np
import tkinter
import tkinter.filedialog
from openpyxl import load_workbook

def select_file():
    """
    Opens dialog box to select source file. Returns string with file path.
    """
    root = tkinter.Tk()
    filename = tkinter.filedialog.askopenfilename(
            title="Open source file",
            filetypes = (("Excel files","*.xlsx"),("all files","*.*")))
    root.destroy()
    return filename

def read_source(source_file,
                sheet,
                num_rows,
                num_cols,
                rows_to_skip=0):
    """
    Wrapper for pandas read_excel function.
    Reads a specific range of cells in an excel file and converts it to a
    numpy array.
    N.B.: it is assumed that all data to be recovered are from column "B"
    of any spreadsheet.

    Parameters:
    source_file       Source excel file.
    sheet             Excel sheet containing desired data range.
    num_rows          Length of data range.
    num_cols          Width of data range.
    rows_to_skip      Nr. of rows to skip before data range.
    """    
    if rows_to_skip == 0:
        skip = []
    else:
        skip = range(rows_to_skip)
    
    df = pd.read_excel(source_file,
                       sheet_name = sheet,
                       header = None,
                       index_col = 0,
                       nrows = num_rows,
                       usecols = range(num_cols+1),
                       skiprows = skip)

    # convert dataframe to numpy array
    var = df.to_numpy()
    var = np.nan_to_num(var, copy=False, nan=-1)
    
    return var

def write_output(output_file,
                 sheet_name,
                 array,
                 idx,
                 cols,
                 rows_to_skip=0,
                 float_format="%.4f"):
    """
    Wrapper for pandas DataFrame.to_excel method.
    Writes a pandas dataframe made of a numpy array, index and column
    titles into a specific excel spreadsheet.

    Parameters:
    output_file       Destination excel file.
    sheet_name        Destination excel sheet.
    array             Numpy array.
    idx               Index for the output dataframe.
    cols              Column titles for the output dataframe.
    rows_to_skip      Nr. of rows to skip before dataframe.
    float_format      Format string for floating point numbers.
                      float_format="%.2f" will format 0.1234 to 0.12.
    """    
    data_df = pd.DataFrame(data = array,
                           index = idx,
                           columns = cols)
    writer = pd.ExcelWriter(output_file, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(output_file)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if rows_to_skip == 0 and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        print("File not found")

    # write out the new sheet
    data_df.to_excel(writer, sheet_name, startrow=rows_to_skip,
                     float_format=float_format)

    # save the workbook
    writer.save()    
    
    


        

    
